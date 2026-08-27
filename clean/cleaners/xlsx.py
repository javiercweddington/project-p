"""
XLSXCleaner - clean Excel files by removing metadata and cleaning content.

Addresses Excel-specific risks (ranked by failure likelihood):

1. Pivot cache (CRITICAL) - pivotCacheRecords retains a FULL copy of source
   data even after the source sheet is deleted. This is ghost content that
   survives visual inspection.

2. Very-hidden sheets (HIGH) - xlSheetVeryHidden sheets don't appear in the
   unhide dialog. Standard location for margin math and sensitive calculations.
   Strategy: Clean their content, then unhide (don't delete - may be referenced).

3. External links (HIGH) - xl/externalLinks/ contains UNC paths to fileservers
   plus cached values pulled from them.

4. Threaded comments (MEDIUM) - persons.xml stores display names and user IDs.

5. Power Query / connections.xml (MEDIUM) - Server names, SQL queries,
   occasionally credentials.

6. Data-validation dropdowns (MEDIUM) - Enumerate your entire vendor list
   even in sheets that look clean.

7. VBA project in .xlsm (MEDIUM) - VBA project passwords are trivially
   bypassed, so treat "protected" as public. Hardcoded paths and developer
   usernames live in modules. Also dump p-code, not just displayed source,
   as they can differ (VBA stomping).

8. Formulas (LOW) - Encode markup multipliers even when displayed values
   look benign. Clean formula text, not just displayed values.

9. BIFF slack in .xls (LOW) - Legacy format retains deleted-cell remnants
   in unused space.

Strategy: Use openpyxl for cell-level cleaning, then directly manipulate
the ZIP internals to remove pivot cache, external links, connections,
and threaded comments at the XML level.

Dependencies (optional):
- openpyxl: Excel file manipulation
"""

from __future__ import annotations

import io
import logging
import os
import re
import shutil
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

from ..anonymizer import EntityMapper
from .text import TextCleaner

_logger = logging.getLogger(__name__)

# Try optional dependencies
try:
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# Fixed ZIP timestamp (MS-DOS epoch: 1980-01-01 00:00:00 UTC)
# Normalizes all zip member dates so the run date doesn't leak.
_FIXED_ZIP_DATE_TIME = (2024, 1, 1, 0, 0, 0)


class XLSXCleaner:
    """Clean Excel files by removing document metadata and cleaning content.

    Uses openpyxl for cell-level cleaning and direct ZIP manipulation
    for XML-level artifact removal.
    """

    # openpyxl core properties that contain identifier-like values
    # and should go through the mapper instead of being blanked.
    # (openpyxl exposes 'creator' and the 'last_modified_by' alias;
    #  Company lives in app.xml, which openpyxl regenerates fresh on save.)
    _IDENTIFIER_PROPERTIES = {
        'creator': ('creator', 'person'),
        'last_modified_by': ('last_modified_by', 'person'),
    }

    # openpyxl core properties that should be blanked (non-identifiers)
    _BLANK_PROPERTIES = {
        'description', 'subject', 'keywords', 'category',
        'title', 'identifier', 'contentStatus',
    }

    # XML paths to remove from the ZIP archive
    _XML_PATHS_TO_REMOVE = {
        # Pivot cache - full copy of source data
        'xl/pivotCache/pivotCacheDefinition1.xml',
        'xl/pivotCache/pivotCacheRecords1.xml',
        # External links - UNC paths to fileservers
        'xl/externalLinks/externalLink1.xml',
        # Connections - server names, SQL, credentials
        'xl/connections.xml',
        # Threaded comments - display names and user IDs
        'xl/threads/persons.xml',
        'xl/threads/threadedComment1.xml',
        # VBA project - hardcoded paths, developer usernames
        'xl/vbaProject.bin',
    }

    # OPC namespaces. NOTE: .rels files use the PACKAGE relationships
    # namespace, not the officeDocument one (which is only for r: attrs).
    CT_NS = '{http://schemas.openxmlformats.org/package/2006/content-types}'
    REL_NS = '{http://schemas.openxmlformats.org/package/2006/relationships}'

    def __init__(self, mapper: EntityMapper):
        self.mapper = mapper
        self.text_cleaner = TextCleaner(mapper)

        if not HAS_OPENPYXL:
            _logger.warning(
                "openpyxl not available; Excel cleaning limited. "
                "Install with: pip install openpyxl"
            )

    def clean_file(self, input_path: Path, output_path: Path) -> bool:
        """Clean an Excel file by removing metadata and cleaning content.

        Args:
            input_path: Source Excel file path
            output_path: Destination Excel file path

        Returns:
            True if cleaning was successful
        """
        ext = input_path.suffix.lower()

        # Legacy .xls (BIFF): no content rewriter, but OLE properties can
        # be zeroed and the raw streams (including BIFF slack remnants)
        # verified free of mapped entities; ship only when that passes.
        if ext == '.xls':
            from .ole_scrub import strip_ole_properties
            return strip_ole_properties(
                self.mapper, input_path, output_path, input_path.name)

        if not HAS_OPENPYXL:
            _logger.warning("openpyxl not available; Excel file fail-closed")
            return False

        try:
            # Load workbook (do NOT preserve VBA - we drop vbaProject.bin)
            wb = load_workbook(input_path, keep_vba=False)

            # Drop external links and pivot caches at the SOURCE so openpyxl
            # never writes the parts, their rels, or workbook.xml
            # <externalReferences> — this is what keeps the output OPC-valid
            # (post-hoc part removal left dangling references).
            try:
                if getattr(wb, '_external_links', None):
                    _logger.info(
                        "Dropping %d external link(s) from %s",
                        len(wb._external_links), input_path.name,
                    )
                    wb._external_links = []
            except Exception as e:
                _logger.debug("Could not drop external links: %s", e)
            for ws in wb.worksheets:
                try:
                    if getattr(ws, '_pivots', None):
                        _logger.info(
                            "Dropping %d pivot table(s) from sheet %r",
                            len(ws._pivots), ws.title,
                        )
                        ws._pivots = []
                except Exception:
                    pass

            # Clear core properties
            self._clear_properties(wb.properties)

            # Clean sheet names (may contain company names); returns the
            # old->new rename map needed to keep formulas consistent
            sheet_renames = self._clean_sheet_names(wb)

            # Clean all cell content (bounded iteration)
            self._clean_all_sheets(wb, sheet_renames)

            # Clean page headers/footers (company name + address routinely
            # sit in every sheet footer) and table names
            for ws in wb.worksheets:
                self._clean_headers_footers(ws)
                self._clean_tables(ws)

            # Handle very-hidden sheets - clean content then unhide
            self._handle_very_hidden_sheets(wb)

            # Save to bytes first (for ZIP manipulation)
            output_bytes = io.BytesIO()
            wb.save(output_bytes)
            output_bytes.seek(0)

            # Post-processing: remove XML-level artifacts by rebuilding ZIP
            output_path.parent.mkdir(parents=True, exist_ok=True)
            self._rebuild_zip_without_artifacts(output_bytes, output_path)

            # Catch-all: entity text in members openpyxl never rewrites
            # (drawings/text boxes/charts). Fail closed if the pass fails.
            from .xml_pass import scrub_zip_xml_members
            if not scrub_zip_xml_members(output_path, self.mapper,
                                         input_path.name):
                output_path.unlink(missing_ok=True)
                return False

            return True

        except Exception as e:
            _logger.error("Error cleaning Excel file %s: %s", input_path, e)
            return False

    def _clean_headers_footers(self, ws) -> None:
        """Clean page header/footer text on a worksheet.

        openpyxl exposes oddHeader/oddFooter/evenHeader/evenFooter/
        firstHeader/firstFooter, each with left/center/right parts.
        """
        for hf_attr in ('oddHeader', 'oddFooter', 'evenHeader', 'evenFooter',
                        'firstHeader', 'firstFooter'):
            hf = getattr(ws, hf_attr, None)
            if hf is None:
                continue
            for side in ('left', 'center', 'right'):
                item = getattr(hf, side, None)
                if item is not None and getattr(item, 'text', None):
                    item.text = self.text_cleaner.clean_text(item.text)

    def _clean_tables(self, ws) -> None:
        """Clean table names, display names, and column names.

        Table (display) names must be bracket-free identifiers, so
        placeholders are inserted with their brackets stripped.
        """
        try:
            tables = list(ws.tables.values())
        except Exception:
            return
        for table in tables:
            for attr in ('name', 'displayName'):
                value = getattr(table, attr, None)
                if value:
                    cleaned = self.text_cleaner.clean_text(value)
                    cleaned = re.sub(r'[\[\]\s]', '_', cleaned)
                    if cleaned != value:
                        try:
                            setattr(table, attr, cleaned)
                        except Exception:
                            pass
            try:
                for column in table.tableColumns:
                    if column.name:
                        column.name = self.text_cleaner.clean_text(column.name)
            except Exception:
                pass

    def _clear_properties(self, props) -> None:
        """Anonymize core document properties.

        Identifier properties (creator, company, etc.) go through the mapper
        to generate consistent placeholders like [PERSON_002].
        Non-identifier properties are blanked.
        Embedded timestamps are normalized to a fixed epoch.
        """
        # Anonymize identifier properties through the mapper
        for prop_name, (_, entity_type) in self._IDENTIFIER_PROPERTIES.items():
            try:
                value = getattr(props, prop_name, None)
                if value and str(value).strip():
                    placeholder = self.mapper.get_or_create(
                        entity_type=entity_type,
                        value=str(value).strip(),
                        source='xlsx_core_properties',
                    )
                    setattr(props, prop_name, placeholder)
                else:
                    setattr(props, prop_name, '')
            except (AttributeError, TypeError):
                pass

        # Blank non-identifier properties
        for prop_name in self._BLANK_PROPERTIES:
            try:
                setattr(props, prop_name, '')
            except (AttributeError, TypeError):
                pass

        # Normalize embedded timestamps to fixed epoch. openpyxl requires
        # datetime objects here; it also re-stamps 'modified' at save time,
        # so the ZIP rebuild pass normalizes docProps/core.xml again.
        fixed_dt = datetime(2024, 1, 1, 0, 0, 0)
        for prop_name in ('created', 'modified', 'lastPrinted'):
            try:
                setattr(props, prop_name, fixed_dt)
            except (AttributeError, TypeError, ValueError):
                pass

    def _clean_sheet_names(self, wb) -> Dict[str, str]:
        """Clean sheet names and defined names to remove identifiers.

        Returns a mapping of {old_title: new_title} so formula references
        to renamed sheets can be kept consistent. Sheet names cannot
        contain []:\\/?* so placeholders are inserted bracket-stripped.
        """
        renames: Dict[str, str] = {}

        sheet_index = 0
        for ws in wb.worksheets:
            sheet_index += 1
            if not ws.title:
                continue
            cleaned = self.text_cleaner.clean_text(ws.title)
            # Excel sheet names have max 31 chars and can't contain :\\/?*[]
            cleaned = re.sub(r'[:\\/?*\[\]]', '', cleaned)
            cleaned = cleaned[:31]

            # Identifying residue: CJK text (unmapped company names) or
            # date-like digit runs make a sheet name identifying even when
            # no mapped entity matched — genericize those wholesale.
            if re.search(r'[぀-ヿ㐀-䶿一-鿿가-힯]|\d{6,}', cleaned):
                generic = f'Sheet{sheet_index}'
                counter = sheet_index
                while generic in wb.sheetnames:
                    counter += 100
                    generic = f'Sheet{counter}'
                cleaned = generic

            # 31-char truncation can collide with an existing sheet; a
            # collision would record the WRONG rename and silently rewire
            # formulas to the wrong sheet.
            if cleaned != ws.title and cleaned in wb.sheetnames:
                suffix = 2
                while f'{cleaned[:28]}_{suffix}' in wb.sheetnames:
                    suffix += 1
                cleaned = f'{cleaned[:28]}_{suffix}'

            if cleaned and cleaned != ws.title:
                renames[ws.title] = cleaned
                ws.title = cleaned

        # Clean defined names (named ranges that may contain paths/names)
        # In modern openpyxl, wb.defined_names is a DefinedNameDict;
        # definedNameFinder was removed in openpyxl 3.1+
        if wb.defined_names:
            # Collect items first since we'll modify the dict during iteration
            defined_items = list(wb.defined_names.items())
            for name, value in defined_items:
                # Propagate sheet renames into the defined-name VALUE
                # ('Old Sheet'!$A$1) or the old sheet name leaks into
                # workbook.xml and the reference dangles.
                try:
                    attr_text = getattr(value, 'attr_text', None) or \
                        getattr(value, 'value', None)
                    if attr_text:
                        new_attr = attr_text
                        for old, new in renames.items():
                            new_attr = new_attr.replace(f"'{old}'!", f"'{new}'!")
                            new_attr = new_attr.replace(f"{old}!", f"{new}!")
                        if new_attr != attr_text:
                            value.value = new_attr
                except Exception:
                    pass

                cleaned_name = self.text_cleaner.clean_text(name)
                # Defined names cannot contain spaces or brackets
                cleaned_name = re.sub(r'[\[\]\s]', '_', cleaned_name)
                if cleaned_name != name:
                    del wb.defined_names[name]
                    # openpyxl 3.1 requires value.name == key on assignment
                    try:
                        value.name = cleaned_name
                    except Exception:
                        pass
                    wb.defined_names[cleaned_name] = value
                    # Record so formulas referencing the defined name
                    # (=Fee_PayPal) are rewritten too — otherwise the old
                    # entity-bearing identifier leaks in every formula and
                    # the reference dangles.
                    renames[name] = cleaned_name

        return renames

    def _clean_all_sheets(self, wb, sheet_renames: Optional[Dict[str, str]] = None) -> None:
        """Clean cell content across all worksheets.

        Iterates only rows that actually exist in the sheet's storage
        (ws._cells / iter_rows over the used range with values_only=False,
        skipping never-written cells) so empty grid intersections are not
        materialized as real cells, which would inflate the output.
        """
        sheet_renames = sheet_renames or {}
        for ws in wb.worksheets:
            # Iterate only cells that exist in the sheet's cell store —
            # ws.iter_rows materializes EmptyCell/Cell objects for every
            # grid intersection in the range, which openpyxl then writes
            # out, inflating the file (1.58MB -> 4.38MB observed).
            existing_cells = list(getattr(ws, '_cells', {}).values())
            if not existing_cells:
                continue

            for cell in existing_cells:
                if cell.value is None:
                    continue

                # ArrayFormula / DataTableFormula objects carry their
                # formula in .text — clean it there (a plain isinstance
                # str-gate silently skipped them).
                if not isinstance(cell.value, str):
                    formula_text = getattr(cell.value, 'text', None)
                    if isinstance(formula_text, str) and formula_text:
                        try:
                            cell.value.text = self._clean_formula_text(
                                formula_text, sheet_renames)
                        except Exception:
                            pass
                    continue

                if isinstance(cell.value, str):
                    if cell.data_type == 'f' or cell.value.startswith('='):
                        # Formula: clean string literals + sheet references
                        cell.value = self._clean_formula_text(
                            cell.value, sheet_renames)
                    else:
                        cell.value = self.text_cleaner.clean_text(cell.value)

                # Clean cell comments (text + author identity)
                if cell.comment:
                    if isinstance(cell.comment.text, str):
                        cell.comment.text = self.text_cleaner.clean_text(
                            cell.comment.text
                        )
                    author = getattr(cell.comment, 'author', None)
                    if author and author.strip():
                        cell.comment.author = self.mapper.get_or_create(
                            'person', author.strip(),
                            source='xlsx_comment_author',
                        )

    def _clean_formula_text(self, formula: str,
                            sheet_renames: Optional[Dict[str, str]] = None) -> str:
        """Clean text within a formula, preserving formula structure.

        In XLSX formulas, single quotes delimit SHEET REFERENCES and double
        quotes delimit STRING LITERALS. Sheet references are rewritten with
        the exact rename map from _clean_sheet_names (bracket-free) so
        references stay valid; string literals get normal entity cleaning.
        """
        sheet_renames = sheet_renames or {}
        if not formula.startswith('='):
            return self.text_cleaner.clean_text(formula)

        result = '='
        rest = formula[1:]

        # Rewrite quoted sheet references: 'Old Sheet Name'!A1
        def _sheet_ref_sub(m: re.Match) -> str:
            ref = m.group(1)
            if ref in sheet_renames:
                return f"'{sheet_renames[ref]}'"
            return m.group(0)
        rest = re.sub(r"'([^']*)'(?=!)", _sheet_ref_sub, rest)

        # Rewrite bare sheet references (OldName!A1) and defined-name
        # identifiers (=Fee_PayPal) alike
        for old, new in sheet_renames.items():
            if re.match(r'^[A-Za-z_][A-Za-z0-9_.]*$', old):
                rest = re.sub(
                    rf'(?<![A-Za-z0-9_.]){re.escape(old)}(?![A-Za-z0-9_.])',
                    new, rest,
                )

        # Clean string literals: VLOOKUP("text", ...)
        rest = re.sub(
            r'"([^"]*)"',
            lambda m: '"' + self.text_cleaner.clean_text(m.group(1)) + '"',
            rest,
        )

        return result + rest

    def _handle_very_hidden_sheets(self, wb) -> None:
        """Handle very-hidden sheets: clean content, then unhide.

        Very-hidden sheets (xlSheetVeryHidden) don't appear in the unhide
        dialog and are a standard location for margin math and sensitive
        calculations. We clean their content and then unhide them so the
        data is visible (and auditable) rather than hidden.
        """
        for ws in wb.worksheets:
            if ws.sheet_state == 'veryHidden':
                _logger.info(
                    "Found very-hidden sheet: %s - cleaning and unhiding",
                    ws.title,
                )
                # Clean the sheet content (already done in _clean_all_sheets)
                # Then unhide
                ws.sheet_state = 'visible'

    def _rebuild_zip_without_artifacts(self, source_bytes: io.BytesIO,
                                        output_path: Path) -> None:
        """Rebuild the Excel ZIP without sensitive XML artifacts.

        This is the correct way to remove files from a ZIP-based format:
        read all entries, skip the ones we want to remove, write a new ZIP.

        Also patches [Content_Types].xml and workbook.xml.rels to maintain
        OPC validity after removing external links and other artifacts.

        Removes:
        - Pivot cache definitions and records
        - External links (with rels patching)
        - Connections
        - Threaded comments (persons.xml)
        - Printer settings
        - VBA project (vbaProject.bin)
        - Normalizes zip member timestamps
        """
        source_bytes.seek(0)

        try:
            # Track which artifact types we remove so we can patch manifests
            removed_external_link_ids = set()
            removed_pivot_cache_defs = set()
            removed_vba = False

            with zipfile.ZipFile(source_bytes, 'r') as src_zf:
                with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as dst_zf:
                    for entry in src_zf.infolist():
                        # Check if this entry should be removed
                        should_remove = False

                        # Check exact paths
                        if entry.filename in self._XML_PATHS_TO_REMOVE:
                            should_remove = True
                            _logger.info("Removing Excel artifact: %s", entry.filename)
                            if entry.filename == 'xl/vbaProject.bin':
                                removed_vba = True

                        # Check patterns (pivot cache, external links)
                        if not should_remove:
                            if 'pivotCache' in entry.filename:
                                should_remove = True
                                _logger.info(
                                    "Removing pivot cache: %s", entry.filename,
                                )
                            elif 'externalLinks' in entry.filename:
                                should_remove = True
                                _logger.info(
                                    "Removing external link: %s", entry.filename,
                                )
                            elif 'threads' in entry.filename:
                                should_remove = True
                                _logger.info(
                                    "Removing threaded comment: %s", entry.filename,
                                )
                            elif 'printerSettings' in entry.filename:
                                should_remove = True
                                _logger.info(
                                    "Removing printer settings: %s", entry.filename,
                                )
                            elif entry.filename == 'xl/connections.xml':
                                should_remove = True
                                _logger.info("Removing connections.xml")

                        if should_remove:
                            continue

                        # Copy entry to new ZIP with normalized timestamp
                        if entry.filename.endswith('/'):
                            # Directory entry
                            new_entry = zipfile.ZipInfo(filename=entry.filename)
                            new_entry.date_time = _FIXED_ZIP_DATE_TIME
                            new_entry.compress_type = zipfile.ZIP_DEFLATED
                            dst_zf.writestr(new_entry, b'')
                        else:
                            data = src_zf.read(entry.filename)

                            # Patch [Content_Types].xml to remove orphaned entries
                            if entry.filename == '[Content_Types].xml':
                                data = self._patch_content_types(
                                    data, removed_vba,
                                )

                            # Patch workbook.xml.rels to remove external link refs
                            if entry.filename == 'xl/_rels/workbook.xml.rels':
                                data = self._patch_workbook_rels(data)

                            # Backup guard: drop <externalReferences> from
                            # workbook.xml if any survived openpyxl
                            if entry.filename == 'xl/workbook.xml':
                                data = re.sub(
                                    rb'<externalReferences>.*?</externalReferences>',
                                    b'', data, flags=re.DOTALL,
                                )

                            # openpyxl re-stamps dcterms:modified with the run
                            # date at save time — normalize it here again.
                            if entry.filename == 'docProps/core.xml':
                                data = re.sub(
                                    rb'(<dcterms:(created|modified)[^>]*>)'
                                    rb'[^<]*(</dcterms:(?:created|modified)>)',
                                    rb'\g<1>2024-01-01T00:00:00Z\g<3>',
                                    data,
                                )

                            # Normalize zip timestamp
                            new_entry = zipfile.ZipInfo(filename=entry.filename)
                            new_entry.date_time = _FIXED_ZIP_DATE_TIME
                            new_entry.compress_type = zipfile.ZIP_DEFLATED
                            new_entry.external_attr = entry.external_attr
                            dst_zf.writestr(new_entry, data)

        except Exception:
            # Fail closed: never ship the unfiltered intermediate bytes.
            if output_path.exists():
                output_path.unlink()
            raise

    def _patch_content_types(self, data: bytes, removed_vba: bool) -> bytes:
        """Patch [Content_Types].xml to remove orphaned content type entries.

        When we remove vbaProject.bin or other artifacts, we must also remove
        the corresponding <Override> entries from [Content_Types].xml, otherwise
        the output is OPC-invalid.
        """
        # Tokens identifying parts this cleaner removes from the package.
        removed_part_tokens = (
            'vbaProject', 'externalLink', 'connections.xml',
            'pivotCache', 'threadedComment', 'persons.xml', 'printerSettings',
        )
        try:
            tree = ET.fromstring(data)

            # NOTE: ElementTree qualified names are '{namespace}Tag' with NO
            # colon — f'{NS}:Override' silently matches nothing.
            for override in list(tree.findall(f'{self.CT_NS}Override')):
                part_name = override.get('PartName') or ''
                content_type = override.get('ContentType', '')
                if (any(tok in part_name for tok in removed_part_tokens)
                        or (removed_vba and 'MacroEnabled' in content_type)):
                    tree.remove(override)
                    _logger.info(
                        "Patched [Content_Types].xml: removed %s", part_name,
                    )

            return ET.tostring(tree, encoding='utf-8', xml_declaration=True)
        except ET.ParseError as e:
            _logger.warning("Failed to parse [Content_Types].xml: %s", e)
            return data

    def _patch_workbook_rels(self, data: bytes) -> bytes:
        """Patch workbook.xml.rels to remove external link relationships.

        Removing externalLinks XML without patching workbook.xml.rels makes
        the output OPC-invalid (Excel repair prompt).
        """
        removed_target_tokens = (
            'externallink', 'connections.xml', 'pivotcache', 'vbaproject',
        )
        try:
            tree = ET.fromstring(data)
            removed_count = 0

            # NOTE: no colon after the namespace (see _patch_content_types).
            for rel in list(tree.findall(f'{self.REL_NS}Relationship')):
                target = (rel.get('Target') or '').lower()
                type_ = (rel.get('Type') or '').lower()
                if any(tok in target or tok in type_
                       for tok in removed_target_tokens):
                    tree.remove(rel)
                    removed_count += 1
                    _logger.info(
                        "Patched workbook.xml.rels: removed rel to %s", target,
                    )

            if removed_count:
                return ET.tostring(tree, encoding='utf-8', xml_declaration=True)
            return data
        except ET.ParseError as e:
            _logger.warning("Failed to parse workbook.xml.rels: %s", e)
            return data

    def detect_risks(self, input_path: Path) -> List[str]:
        """Detect potential risk vectors in an Excel file.

        Args:
            input_path: Source Excel file path

        Returns:
            List of detected risk descriptions
        """
        risks = []
        ext = input_path.suffix.lower()

        if ext == '.xls':
            risks.append(
                "Legacy BIFF format - deleted-cell remnants may persist in slack space"
            )

        if ext == '.xlsm':
            risks.append(
                "Macro-enabled workbook - VBA project passwords are trivially bypassed. "
                "Treat 'protected' VBA as public."
            )

        if not zipfile.is_zipfile(input_path):
            return risks

        try:
            with zipfile.ZipFile(input_path, 'r') as zf:
                namelist = zf.namelist()

                # Check for pivot cache
                pivot_files = [f for f in namelist if 'pivotCache' in f]
                if pivot_files:
                    risks.append(
                        f"Pivot cache detected ({len(pivot_files)} files) - "
                        f"full copy of source data may persist even after source sheet deletion"
                    )

                # Check for external links
                ext_link_files = [f for f in namelist if 'externalLinks' in f]
                if ext_link_files:
                    risks.append(
                        f"External links detected ({len(ext_link_files)} files) - "
                        f"may contain UNC paths to fileservers"
                    )

                # Check for connections
                if 'xl/connections.xml' in namelist:
                    risks.append(
                        "Connections.xml present - may contain server names, SQL queries, credentials"
                    )

                # Check for threaded comments
                thread_files = [f for f in namelist if 'threads' in f]
                if thread_files:
                    risks.append(
                        f"Threaded comments present ({len(thread_files)} files) - "
                        f"display names and user IDs in persons.xml"
                    )

                # Check for very-hidden sheets
                try:
                    workbooks_xml = zf.read('xl/workbook.xml').decode('utf-8')
                    if 'veryHidden' in workbooks_xml:
                        risks.append(
                            "Very-hidden sheets detected - may contain margin math or sensitive data"
                        )
                except Exception:
                    pass

                # Check for data validation (vendor lists)
                try:
                    for sheet_path in namelist:
                        if sheet_path.startswith('xl/worksheets/sheet'):
                            sheet_xml = zf.read(sheet_path).decode('utf-8')
                            if 'dataValidation' in sheet_xml:
                                risks.append(
                                    f"Data validation in {sheet_path} - may enumerate vendor lists"
                                )
                                break
                except Exception:
                    pass

                # Check for VBA project
                if 'xl/vbaProject.bin' in namelist:
                    risks.append(
                        "VBA project present - may contain hardcoded paths, developer usernames"
                    )

        except Exception as e:
            risks.append(f"Risk detection failed: {e}")

        return risks